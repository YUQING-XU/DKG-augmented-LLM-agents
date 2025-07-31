import ifcopenshell
import csv
from collections import defaultdict
import openpyxl

def get_property_value(properties, property_name):
    if properties:
        for prop in properties:
            if prop.Name == property_name:
                return prop.NominalValue.wrappedValue
    return None

def get_pset_value(element, pset_name, property_name):
    for definition in element.IsDefinedBy:
        if definition.is_a("IfcRelDefinesByProperties"):
            property_set = definition.RelatingPropertyDefinition
            if property_set.is_a("IfcPropertySet") and property_set.Name == pset_name:
                for prop in property_set.HasProperties:
                    if prop.Name == property_name:
                        if hasattr(prop, "NominalValue"):
                            return prop.NominalValue.wrappedValue
    return None

def get_type_pset_value(element_type, pset_name, property_name):
    # Loop through all property sets of the type element
    for definition in element_type.HasPropertySets:
        # Check if it's the property set we're looking for
        if definition.is_a("IfcPropertySet") and definition.Name == pset_name:
            # Loop through properties in the property set
            for prop in definition.HasProperties:
                if prop.Name == property_name:
                    if hasattr(prop, "NominalValue"):
                        return prop.NominalValue.wrappedValue
                    
def get_element_type(element):
    # Get the type object of the element
    for definition in element.IsDefinedBy:
        if definition.is_a("IfcRelDefinesByType"):
            return definition.RelatingType
    return None

def extract_air_terminals(ifc_file_path):
    # Load IFC file
    ifc_file = ifcopenshell.open(ifc_file_path)
    
    # Get all air terminals
    air_terminals = ifc_file.by_type('IfcAirTerminal')
    
    # Organize by family type
    terminals_by_type = defaultdict(list)
    
    for terminal in air_terminals:
        # Get basic info
        terminal_data = {
            'GlobalId': terminal.GlobalId,
            'Name': terminal.Name,
            'Type': terminal.is_a(),
            'Size': None,
            'FlowRate': None,
            'Manufacturer': None,
            'MaxFlow': None,
            'NC': None,
            'PropertySets': {}
        }
        
        # Get family type (IfcTypeObject)
        if terminal.IsTypedBy:
            type_obj = terminal.IsTypedBy[0].RelatingType
            terminal_data['FamilyType'] = type_obj.Name or type_obj.is_a()
            
            # Extract Manufacturer and NC from IfcAirTerminalType
            if type_obj.is_a('IfcAirTerminalType'):
                # Access property sets through the correct relationship
                for definition in type_obj.HasPropertySets:
                    if definition.is_a('IfcPropertySet'):
                        for p in definition.HasProperties:
                            value = None
                            if p.is_a('IfcPropertySingleValue'):
                                value = p.NominalValue.wrappedValue if p.NominalValue else None
                            elif p.is_a('IfcPropertyEnumeratedValue'):
                                value = [v.wrappedValue for v in p.EnumerationValues] if p.EnumerationValues else None
                            
                            if p.Name == 'Manufacturer':
                                terminal_data['Manufacturer'] = value
                            if p.Name == 'NC':
                                terminal_data['NC'] = value
        else:
            terminal_data['FamilyType'] = "Ungrouped"
        
        # Extract properties from both the instance and its type
        for definition in terminal.IsDefinedBy:
            if definition.is_a('IfcRelDefinesByProperties'):
                prop_set = definition.RelatingPropertyDefinition
                if prop_set.is_a('IfcPropertySet'):
                    terminal_data['PropertySets'][prop_set.Name] = {
                        p.Name: (p.NominalValue.wrappedValue if p.is_a('IfcPropertySingleValue') and p.NominalValue else None)
                        for p in prop_set.HasProperties
                    }
        
        # Try to get size and flow from common property sets
        for prop_set_name, props in terminal_data['PropertySets'].items():
            # Size extraction
            if not terminal_data['Size']:
                for size_prop in ['Size', 'NominalSize', 'OverallSize', 'Dimensions']:
                    if size_prop in props:
                        terminal_data['Size'] = props[size_prop]
                        break
            
            # Flow rate extraction
            if not terminal_data['FlowRate']:
                for flow_prop in ['FlowRate', 'AirFlowRate', 'NominalFlowRate', 'Capacity']:
                    if flow_prop in props:
                        terminal_data['FlowRate'] = props[flow_prop]
                        break
        
        # Add to our organized dictionary
        terminals_by_type[terminal_data['FamilyType']].append(terminal_data)
    
    return terminals_by_type

def extract_duct_fittings(ifc_file_path):
    # Load IFC file
    ifc_file = ifcopenshell.open(ifc_file_path)
    
    # Get all air terminals
    duct_fittings = ifc_file.by_type('IfcDuctFitting')
    
    # Organize by family type
    duct_fittings_by_type = defaultdict(list)
    
    for duct_fitting in duct_fittings:
        # Get basic info
        duct_fitting_data = {
            'GlobalId': duct_fitting.GlobalId,
            'Name': duct_fitting.Name,
            'Type': duct_fitting.is_a(),
            'Size': None,
            'Volume': None,
            'Material': None,
            'Manufacturer': None,
            'EmissionFactor': None,
            'PropertySets': {}
        }
        
        # Get family type (IfcTypeObject)
        if duct_fitting.IsTypedBy:
            type_obj = duct_fitting.IsTypedBy[0].RelatingType
            duct_fitting_data['FamilyType'] = type_obj.Name or type_obj.is_a()
            
            # Extract Manufacturer from IfcDuctFittingType
            if type_obj.is_a('IfcDuctFittingType'):
                # Access property sets through the correct relationship
                for definition in type_obj.HasPropertySets:
                    if definition.is_a('IfcPropertySet'):
                        for p in definition.HasProperties:
                            value = None
                            if p.is_a('IfcPropertySingleValue'):
                                value = p.NominalValue.wrappedValue if p.NominalValue else None
                            elif p.is_a('IfcPropertyEnumeratedValue'):
                                value = [v.wrappedValue for v in p.EnumerationValues] if p.EnumerationValues else None
                            
                            if p.Name == 'Manufacturer':
                                duct_fitting_data['Manufacturer'] = value
                            if p.Name == 'Material':
                                duct_fitting_data['Material'] = value
        else:
            duct_fitting_data['FamilyType'] = "Ungrouped"
        
        # Extract properties from both the instance and its type
        for definition in duct_fitting.IsDefinedBy:
            if definition.is_a('IfcRelDefinesByProperties'):
                prop_set = definition.RelatingPropertyDefinition
                if prop_set.is_a('IfcPropertySet'):
                    duct_fitting_data['PropertySets'][prop_set.Name] = {
                        p.Name: (p.NominalValue.wrappedValue if p.is_a('IfcPropertySingleValue') and p.NominalValue else None)
                        for p in prop_set.HasProperties
                    }
        
        # Try to get size and flow from common property sets
        for prop_set_name, props in duct_fitting_data['PropertySets'].items():
            # Size extraction
            if not duct_fitting_data['Size']:
                for size_prop in ['Size', 'NominalSize', 'OverallSize', 'Dimensions']:
                    if size_prop in props:
                        duct_fitting_data['Size'] = props[size_prop]
                        break
            
            # Volume extraction
            if not duct_fitting_data['Volume']:
                for volume_prop in ['Volume', 'NominalVolume', 'OverallVolume', 'Dimensions']:
                    if volume_prop in props:
                        duct_fitting_data['Volume'] = props[volume_prop]
                        break
            
            # Emission factor extraction
            if not duct_fitting_data['EmissionFactor']:
                for emission_prop in ['EmissionFactor', 'ClimateChangePerUnit']:
                    if emission_prop in props:
                        duct_fitting_data['EmissionFactor'] = props[emission_prop]
                        break
        
        # Add to our organized dictionary
        duct_fittings_by_type[duct_fitting_data['FamilyType']].append(duct_fitting_data)
    
    return duct_fittings_by_type


def extract_duct_segments(ifc_file_path):
    # Load IFC file
    ifc_file = ifcopenshell.open(ifc_file_path)
    
    # Get all air terminals
    duct_segments = ifc_file.by_type('IfcDuctSegment')
    
    # Organize by family type
    duct_segments_by_type = defaultdict(list)
    
    for duct in duct_segments:
        # Get basic info
        duct_data = {
            'GlobalId': duct.GlobalId,
            'Name': duct.Name,
            'Type': duct.is_a(),
            'Material': None,
            'Thickness': None,
            'Size': None,
            'Length': None,
            'Width': None,
            'Height': None,
            'RelativeRoughness': None,
            'FlowRate': None,
            'Manufacturer': None,
            'EmissionFactor': None,
            'PropertySets': {}
        }
        
        # Get family type (IfcTypeObject)
        if duct.IsTypedBy:
            type_obj = duct.IsTypedBy[0].RelatingType
            duct_data['FamilyType'] = type_obj.Name or type_obj.is_a()
            
            # Extract Manufacturer, material, and thickness from IfcDuctSegmentType
            if type_obj.is_a('IfcDuctSegmentType'):
                # Access property sets through the correct relationship
                for definition in type_obj.HasPropertySets:
                    if definition.is_a('IfcPropertySet'):
                        for p in definition.HasProperties:
                            value = None
                            if p.is_a('IfcPropertySingleValue'):
                                value = p.NominalValue.wrappedValue if p.NominalValue else None
                            elif p.is_a('IfcPropertyEnumeratedValue'):
                                value = [v.wrappedValue for v in p.EnumerationValues] if p.EnumerationValues else None
                            
                            if p.Name == 'Manufacturer':
                                duct_data['Manufacturer'] = value
                            if p.Name == 'Material':
                                duct_data['Material'] = value
                            if p.Name == 'Thickness':
                                duct_data['Thickness'] = value
        else:
            duct_data['FamilyType'] = "Ungrouped"
        
        # Extract properties from both the instance and its type
        for definition in duct.IsDefinedBy:
            if definition.is_a('IfcRelDefinesByProperties'):
                prop_set = definition.RelatingPropertyDefinition
                if prop_set.is_a('IfcPropertySet'):
                    duct_data['PropertySets'][prop_set.Name] = {
                        p.Name: (p.NominalValue.wrappedValue if p.is_a('IfcPropertySingleValue') and p.NominalValue else None)
                        for p in prop_set.HasProperties
                    }
                if prop_set.Name == 'Mechanical - Flow':
                    for prop in prop_set.HasProperties:
                        if prop.Name == 'Relative Roughness':
                            duct_data['RelativeRoughness'] = prop.NominalValue.wrappedValue if prop.is_a('IfcPropertySingleValue') and prop.NominalValue else None
                            break
                # Extract ClimateChangePerUnit
                if prop_set.Name == 'Pset_EnvironmentalImpactIndicators':
                    for prop in prop_set.HasProperties:
                        if prop.Name == 'ClimateChangePerUnit':
                            duct_data['EmissionFactor'] = prop.NominalValue.wrappedValue 
                            break

        # Try to get size and flow from common property sets
        for prop_set_name, props in duct_data['PropertySets'].items():
            # Size extraction
            if not duct_data['Size']:
                for size_prop in ['Size', 'NominalSize', 'OverallSize', 'Dimensions']:
                    if size_prop in props:
                        duct_data['Size'] = props[size_prop]
                        break
            # Length extraction
            if not duct_data['Length']:
                for length_prop in ['Length', 'OverallLength']:
                    if length_prop in props:
                        duct_data['Length'] = props[length_prop]
                        break
            # Width extraction
            if not duct_data['Width']:
                for width_prop in ['Width', 'OverallWidth']:
                    if width_prop in props:
                        duct_data['Width'] = props[width_prop]
                        break
            # Height extraction
            if not duct_data['Height']:
                for height_prop in ['Height', 'OverallHeight']:
                    if height_prop in props:
                        duct_data['Height'] = props[height_prop]
                        break
      
            # Flow rate extraction
            if not duct_data['FlowRate']:
                for flow_prop in ['FlowRate', 'AirFlowRate', 'NominalFlowRate', 'Capacity']:
                    if flow_prop in props:
                        duct_data['FlowRate'] = props[flow_prop]
                        break
            # Emission factor extraction
            # if not duct_data['EmissionFactor']:
            #     for emission_prop in ['ClimateChangePerUnit']:
            #         if emission_prop in props:
            #             duct_data['EmissionFactor'] = props[emission_prop]
            #             break
        
        # Add to our organized dictionary
        duct_segments_by_type[duct_data['FamilyType']].append(duct_data)
    
    return duct_segments_by_type


def save_to_AirTerminal_excel(terminals_by_type, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Air Terminals"
    
    # Define headers
    headers = ['FamilyType', 'GlobalId', 'Name', 'Type', 'Size', 'FlowRate', 'Manufacturer', 'MaxFlow', 'NC']
    sheet.append(headers)
    
    # Write data
    for family_type, terminals in terminals_by_type.items():
        for terminal in terminals:
            sheet.append([
                family_type,
                terminal['GlobalId'],
                terminal['Name'],
                terminal['Type'],
                terminal['Size'],
                terminal['FlowRate'],
                terminal['Manufacturer'],
                terminal['MaxFlow'],
                terminal['NC']
            ])
    
    workbook.save(output_file)

def save_to_DuctSegment_excel(duct_segments_by_type, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Duct Segments"
    
    # Define headers
    headers = ['FamilyType', 'GlobalId', 'Name', 'Type', 'Material', 'Thickness', 'Size', 'Length', 'Width', 'Height', 'RelativeRoughness', 'FlowRate', 'Manufacturer', 'EmissionFactor']
    sheet.append(headers)
    
    # Write data
    for family_type, ducts in duct_segments_by_type.items():
        for duct in ducts:
            sheet.append([
                family_type,
                duct['GlobalId'],
                duct['Name'],
                duct['Type'],
                duct['Material'],
                duct['Thickness'],
                duct['Size'],
                duct['Length'],
                duct['Width'],
                duct['Height'],
                duct['RelativeRoughness'],
                duct['FlowRate'],
                duct['Manufacturer'],
                duct['EmissionFactor']
            ])
    
    workbook.save(output_file)

def save_to_DuctFitting_excel(duct_fittings_by_type, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Duct Fittings"
    
    # Define headers
    headers = ['FamilyType', 'GlobalId', 'Name', 'Type', 'Size', 'Volume', 'Material', 'Manufacturer', 'EmissionFactor']
    sheet.append(headers)
    
    # Write data
    for family_type, ducts in duct_fittings_by_type.items():
        for duct in ducts:
            sheet.append([
                family_type,
                duct['GlobalId'],
                duct['Name'],
                duct['Type'],
                duct['Size'],
                duct['Volume'],
                duct['Material'],
                duct['Manufacturer'],
                duct['EmissionFactor']
            ])
    
    workbook.save(output_file)

# Usage
ifc_path = 'sampleMEP/IFC/DuctSystems-2024-0409-2.ifc'
air_terminals_output_excel = 'sampleMEP/air_terminals_output.xlsx'
duct_segments_output_excel = 'sampleMEP/duct_segments_output.xlsx'
duct_fittings_output_excel = 'sampleMEP/duct_fittings_output.xlsx'

terminals_data = extract_air_terminals(ifc_path)
save_to_AirTerminal_excel(terminals_data, air_terminals_output_excel)

duct_segments_data = extract_duct_segments(ifc_path)
save_to_DuctSegment_excel(duct_segments_data, duct_segments_output_excel)

duct_fittings_data = extract_duct_fittings(ifc_path)
save_to_DuctFitting_excel(duct_fittings_data, duct_fittings_output_excel)

print(f"Successfully extracted {sum(len(v) for v in terminals_data.values())} air terminals to {air_terminals_output_excel}")
print(f"Successfully extracted {sum(len(v) for v in duct_segments_data.values())} duct segments to {duct_segments_output_excel}")
print(f"Successfully extracted {sum(len(v) for v in duct_fittings_data.values())} duct fittings to {duct_fittings_output_excel}")